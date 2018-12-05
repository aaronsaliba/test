import openpyxl as xl
import openpyxl.styles as xlst
import requests as req
import datetime as dt
import pandas as pd
import bs4

def today_in_cell(workbook, sheet, cell):
    # today_in_cell("testing.xlsx","Table","G4")
    """[Writes today's date and time in a cell]
    
    Arguments:
        workbook {[string]} -- [relative path of excel sheet. ex. "abc.xlsx"]
        sheet {[string]} -- [name of excel sheet. ex. "abc"]
        cell {[string]} -- [cell reference. ex: "A1"]
    """
    
    wb = xl.load_workbook(filename=workbook)
    ws = wb[sheet] 
    ws[cell].number_format
    'dd/mm/yyyy hh:mm:ss'
    ws[cell] = dt.datetime.today()
    wb.save(workbook)

def string_in_cell(workbook, sheet, cell, string):
    # string_in_cell("testing.xlsx","Table","B4","")
    """[Writes string in cell]
    
    Arguments:
        workbook {[string]} -- [relative path of excel sheet. ex. "abc.xlsx"]
        sheet {[string]} -- [name of excel sheet. ex. "abc"]
        cell {[string]} -- [cell reference. ex: "A1"]
        string {[string]} -- [string to write in cell]
    """
    
    wb = xl.load_workbook(filename=workbook)
    ws = wb[sheet]
    ws[cell] = string
    wb.save(workbook)

def cell_reader(workbook, sheet, cell):
    
    """[Reads a cell's value]
    
    Arguments:
        workbook {[string]} -- [relative path of excel sheet. ex. "abc.xlsx"]
        sheet {[string]} -- [name of excel sheet. ex. "abc"]
        cell {[string]} -- [cell reference. ex: "A1"]
    """
    wb = xl.load_workbook(filename=workbook)
    ws = wb[sheet]
    val = ws[cell].value
    return val

def load_soy_data():
        
    """[loads the begining of the year data in excel]
        
    Arguments:
        workbook {[string]} -- [relative path of excel sheet. ex. "abc.xlsx"]
        sheet {[string]} -- [name of excel sheet. ex. "abc"]
    """
    todays_date = dt.datetime.today()
    current_year = todays_date.year
    url = f"https://coinmarketcap.com/historical/20131001/"
    data =req.get(url).text
    soup = bs4.BeautifulSoup(data,"lxml")
    table = soup.find(id ="currencies-all" )
    tbody = table.find('tbody')
    allTrs = tbody.find_all('tr')
    for tr in allTrs:
        for td in tr.find('td'):
            t = td


    
load_soy_data()