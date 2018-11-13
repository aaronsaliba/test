import openpyxl as xl
import openpyxl.styles as xlst
import requests as req
import datetime as dt

class vars:
    def __init__(self, work_book, work_sheet, cell):

        """[3 essencial variables for populating a cell]
        
        Arguments:
            work_book {[string]} -- [the name of the workbook]
            work_sheet {[string]} -- [the name of the worksheet]
            cell {[string]} -- [the name of the cell]
        """

        self.work_book = work_book
        self.work_sheet = work_sheet
        self.cell = cell
    def active_testing_sheet(self):
        wb = xl.load_workbook(self.work_book)
        wb.save(self.work_book)
    def today_in_cell(self):
        wb = xl.load_workbook(self.work_book)
        ws = wb[self.work_sheet]
        ws[self.cell].number_format
        'dd/mm/yyyy'
        ws[self.cell] = dt.datetime.today()
        wb.save(self.work_book)


ws = vars("testing.xlsx","Table","G4")
ws.active_testing_sheet()
ws.today_in_cell()
