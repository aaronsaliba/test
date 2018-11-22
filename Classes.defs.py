import openpyxl as xl
import openpyxl.styles as xlst
import requests as req
import datetime as dt

def today_in_cell(workbook, sheet, cell):
    """[Writes today's date and time in a cell]
    
    Arguments:
        workbook {[string]} -- [relative path of excel sheet. ex. "abc.xlsx" ]
        sheet {[string]} -- [name of excel sheet. ex. "abc"]
        cell {[string]} -- [cell reference. ex: "A1"]
    """
    
    wb = xl.load_workbook(filename=workbook)
    ws = wb[sheet]
    ws[cell].number_format
    'dd/mm/yyyy hh:mm:ss'
    ws[cell] = dt.datetime.today()
    wb.save(workbook)
def write_string_in_cell(workbook, sheet, cell):
    """[Writes a string in a cell]
    
    Arguments:
        workbook {[string]} -- [description]
        sheet {[string]} -- [description]
        cell {[string]} -- [description]
    """

    wb = xl.load_workbook(filename=workbook)
    ws = wb[sheet]
    ws[cell].number_format
    'dd/mm/yyyy h:mm:ss'
    ws[cell] = dt.datetime.today()
    wb.save(workbook)
    

today_in_cell("testing.xlsx","Table","G4")